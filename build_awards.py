# -*- coding: utf-8 -*-
"""
build_awards.py — Regenera Awards.html a partir de los Excel en /Awards.

Uso:
    python build_awards.py            # revisa cambios y reconstruye solo si es necesario
    python build_awards.py --force    # reconstruye siempre, aunque no detecte cambios

Pensado para correr como Tarea Programada de Windows cada cierto tiempo
(ver setup_scheduled_task.ps1). Si algún Excel está abierto/bloqueado en ese
momento, se salta con una advertencia en el log y se reintenta en la próxima
corrida — nunca deja Awards.html a medio escribir.
"""
import io
import json
import os
import re
import sys
import traceback
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
AWARDS_DIR = os.path.join(BASE_DIR, 'Awards')
INDEX_HTML = os.path.join(BASE_DIR, 'index.html')
TEMPLATE_HTML = os.path.join(BASE_DIR, '_awards_template.html')
OUTPUT_HTML = os.path.join(BASE_DIR, 'Awards.html')
STATE_FILE = os.path.join(BASE_DIR, '_awards_build_state.json')
LOG_FILE = os.path.join(BASE_DIR, 'awards_build.log')
SALES_REPORT_FILE = os.path.join(BASE_DIR, 'Informe de ventas (Real+1).xlsx')

WEEKDAYS_ES = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
MONTHS_ES = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio',
             'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']

SOURCE_FILES = {
    'arboleda':  os.path.join(AWARDS_DIR, 'Awards ARBOLEDA.xlsx'),
    'caliterra': os.path.join(AWARDS_DIR, 'Awards CALITERRA.xlsx'),
    'errazuriz': os.path.join(AWARDS_DIR, 'Awards ERRAZURIZ_NEW.xlsx'),
    'sena_vch':  os.path.join(AWARDS_DIR, 'Awards SÑ - VCH.xls'),
}


def log(msg):
    line = f"[{datetime.now().isoformat(timespec='seconds')}] {msg}"
    print(line)
    with io.open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(line + '\n')


def load_state():
    if os.path.exists(STATE_FILE):
        try:
            with io.open(STATE_FILE, encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_state(state):
    with io.open(STATE_FILE, 'w', encoding='utf-8') as f:
        json.dump(state, f, indent=2)


def current_mtimes():
    mtimes = {}
    for key, path in SOURCE_FILES.items():
        if os.path.exists(path):
            mtimes[key] = os.path.getmtime(path)
    return mtimes


def needs_rebuild(force=False):
    if force:
        return True
    old = load_state().get('mtimes', {})
    new = current_mtimes()
    if old != new:
        return True
    if not os.path.exists(OUTPUT_HTML):
        return True
    return False


def clean(v):
    if v is None:
        return ''
    if isinstance(v, str):
        return v.strip()
    return v


def extract_data():
    import openpyxl
    import xlrd

    records = []

    def add(vina, submarca, categoria, cepa, cosecha, concurso, puntaje, premio, anio):
        if not any([submarca, cepa, concurso, puntaje, premio]):
            return
        records.append({
            'vina': vina, 'submarca': clean(submarca) or '—', 'categoria': clean(categoria) or '',
            'cepa': clean(cepa) or '—', 'cosecha': clean(cosecha) or '—',
            'concurso': clean(concurso) or '—', 'puntaje': clean(puntaje) or '—',
            'premio': clean(premio) or '—', 'anio': clean(anio) or ''
        })

    # ── ARBOLEDA ──────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(SOURCE_FILES['arboleda'], data_only=True)
    ws = wb['ARBOLEDA']
    for row in ws.iter_rows(min_row=2, max_row=500, values_only=True):
        if not row or not row[0]:
            continue
        _, categoria, submarca, concurso, mercado, cepa, cosecha, puntaje, premio, fecha, anio, *_ = row + (None,) * 20
        add('Arboleda', submarca, categoria, cepa, cosecha, concurso, puntaje, premio, anio)

    # ── CALITERRA ─────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(SOURCE_FILES['caliterra'], data_only=True)
    ws = wb['CALITERRA']
    for row in ws.iter_rows(min_row=2, max_row=560, values_only=True):
        if not row or not row[0]:
            continue
        _, categoria, anio1, fenvio, concurso, mercado, submarca, variedad, cosecha, puntaje, premio, fpubl, anio2, *_ = row + (None,) * 20
        add('Caliterra', submarca, categoria, variedad, cosecha, concurso, puntaje, premio, anio2 or anio1)

    # ── ERRÁZURIZ ─────────────────────────────────────────────────────────
    # Excluye filas con STATUS = "Discontinuado" y categoría "Tactica".
    wb = openpyxl.load_workbook(SOURCE_FILES['errazuriz'], data_only=True)
    ws = wb['ERRAZURIZ']
    skipped_disc = 0
    skipped_tactica = 0
    for row in ws.iter_rows(min_row=2, max_row=2200, values_only=True):
        if not row or not row[0]:
            continue
        (_, categoria, submarca, cepa, cosecha, premiador, puntaje, premio, fecha, anio,
         pais, comentario, nivel, apelacion, status, *_rest) = row + (None,) * 20
        status_norm = (status or '').strip().lower() if isinstance(status, str) else ''
        if status_norm == 'discontinuado':
            skipped_disc += 1
            continue
        cat_norm = (categoria or '').strip().lower()
        if cat_norm == 'tactica':
            skipped_tactica += 1
            continue
        add('Errázuriz', submarca, categoria, cepa, cosecha, premiador, puntaje, premio, anio)
    log(f"Errázuriz: omitidas {skipped_disc} filas Discontinuado, {skipped_tactica} filas Tactica")

    # ── SEÑA / VIÑEDO CHADWICK ────────────────────────────────────────────
    wbx = xlrd.open_workbook(SOURCE_FILES['sena_vch'])
    for sheet_idx, vina_label in [(0, 'Seña'), (1, 'Viñedo Chadwick')]:
        ws = wbx.sheet_by_index(sheet_idx)
        for r in range(1, ws.nrows):
            row = ws.row_values(r)
            if not row or not row[0]:
                continue
            wine, vintage, source, score, premio, idate, iyear = (row + [None] * 8)[:7]
            vintage = int(vintage) if isinstance(vintage, (int, float)) and vintage else vintage
            iyear = int(iyear) if isinstance(iyear, (int, float)) and iyear else iyear
            add(vina_label, vina_label, '', '—', vintage, source, score, premio, iyear)

    return records


def extract_logo_line():
    """Saca la línea <img> del header (logo VFCh) directamente de index.html,
    para no depender de una copia separada que podría quedar desactualizada."""
    with io.open(INDEX_HTML, encoding='utf-8') as f:
        for line in f:
            if '<img src="data:image' in line:
                return line.rstrip('\n')
    raise RuntimeError('No se encontró la línea del logo en index.html')


def build(force=False):
    if not needs_rebuild(force=force):
        log('Sin cambios en los Excel — no se reconstruye.')
        return False

    try:
        records = extract_data()
    except PermissionError as e:
        log(f'ADVERTENCIA: un Excel está abierto/bloqueado, se reintentará en la próxima corrida. ({e})')
        return False
    except Exception:
        log('ERROR extrayendo datos de los Excel:\n' + traceback.format_exc())
        return False

    with io.open(TEMPLATE_HTML, encoding='utf-8') as f:
        template = f.read()

    logo_line = extract_logo_line()
    data_json = json.dumps(records, ensure_ascii=False)

    out = template.replace('__LOGO_IMG__', logo_line)
    out = out.replace('__AWARDS_DATA__', data_json)

    tmp_path = OUTPUT_HTML + '.tmp'
    with io.open(tmp_path, 'w', encoding='utf-8') as f:
        f.write(out)
    os.replace(tmp_path, OUTPUT_HTML)  # escritura atómica: nunca deja el html a medio escribir

    save_state({'mtimes': current_mtimes(), 'last_build': datetime.now().isoformat(), 'records': len(records)})
    log(f'Awards.html reconstruido correctamente — {len(records)} registros.')
    return True


def format_spanish_date(dt):
    weekday = WEEKDAYS_ES[dt.weekday()]
    month = MONTHS_ES[dt.month - 1].capitalize()
    return f"{weekday} {dt.day} de {month} {dt.year}"


def update_sales_report_label():
    """Actualiza el texto "Actualizado por última vez el ..." del botón
    INFORME DE VENTAS según la fecha real de modificación del Excel.
    El link del botón NO se toca (queda fijo, ya que Cinthia sobrescribe
    siempre el mismo archivo, así que el link para compartir de OneDrive
    no cambia)."""
    if not os.path.exists(SALES_REPORT_FILE):
        log('ADVERTENCIA: no se encontró "Informe de ventas (Real+1).xlsx" para actualizar la fecha.')
        return

    mtime = os.path.getmtime(SALES_REPORT_FILE)
    dt = datetime.fromtimestamp(mtime)
    new_label = f"Actualizado por última vez el {format_spanish_date(dt)}"

    try:
        with io.open(INDEX_HTML, encoding='utf-8') as f:
            content = f.read()
    except PermissionError:
        log('ADVERTENCIA: index.html está abierto/bloqueado, se reintentará en la próxima corrida.')
        return

    pattern = re.compile(r'(<span class="tool-btn-sub">)Actualizado por última vez el [^<]*(</span>)')
    match = pattern.search(content)
    if not match:
        log('ADVERTENCIA: no se encontró el texto "Actualizado por última vez" en index.html.')
        return

    if new_label in match.group(0):
        log('Informe de Ventas: fecha sin cambios, no se toca index.html.')
        return

    new_content = pattern.sub(lambda m: m.group(1) + new_label + m.group(2), content, count=1)

    try:
        tmp_path = INDEX_HTML + '.tmp'
        with io.open(tmp_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        os.replace(tmp_path, INDEX_HTML)
    except PermissionError:
        log('ADVERTENCIA: no se pudo escribir index.html (bloqueado), se reintentará en la próxima corrida.')
        return

    log(f'index.html actualizado: "{new_label}"')


if __name__ == '__main__':
    force = '--force' in sys.argv
    try:
        update_sales_report_label()
    except Exception:
        log('ERROR actualizando fecha del Informe de Ventas:\n' + traceback.format_exc())
    try:
        build(force=force)
    except Exception:
        log('ERROR inesperado:\n' + traceback.format_exc())
        sys.exit(1)
